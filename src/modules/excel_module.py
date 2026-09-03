import os
from src.core.base_module import BaseDocumentModule
from src.core.registry import ModuleRegistry

class ExcelModule(BaseDocumentModule):
    @property
    def name(self) -> str:
        return "Excel"

    @property
    def file_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]

    @property
    def required_dependencies(self) -> list[str]:
        return ["openpyxl"]

    def load_to_markdown(self, file_path: str) -> str:
        """Extracts Excel sheets into clean Markdown tables, preserving bold, italic, strike, underline, and hyperlinks."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            import openpyxl
            from src.core.converters import wrap_text_style
            from openpyxl.cell.rich_text import CellRichText

            wb = openpyxl.load_workbook(file_path, data_only=True, rich_text=True)
            parts = []
            for name in wb.sheetnames:
                ws = wb[name]
                parts.append(f"## {name}\n")
                
                rows = list(ws.iter_rows(values_only=False))
                if not rows:
                    parts.append("*(Empty Table)*\n")
                    continue
                    
                grid = []
                for row in rows:
                    row_cells = []
                    for cell in row:
                        val = cell.value
                        val_str = ""
                        
                        if isinstance(val, CellRichText):
                            formatted_parts = []
                            for el in val:
                                if isinstance(el, str):
                                    formatted_parts.append(el)
                                else:
                                    font = el.font
                                    bold = font.b if font else False
                                    italic = font.i if font else False
                                    strike = font.strike if font else False
                                    underline = bool(font.u) if font and font.u else False
                                    code = font.rFont == "Consolas" if font and font.rFont else False
                                    formatted_parts.append(wrap_text_style(el.text, bold=bold, italic=italic, strike=strike, underline=underline, code=code))
                            val_str = "".join(formatted_parts).strip()
                        else:
                            val_str = str(val).strip() if val is not None else ""
                            if val_str and cell.has_style:
                                font = cell.font
                                bold = font.bold if font else False
                                italic = font.italic if font else False
                                strike = font.strike if font else False
                                underline = bool(font.underline) if font and font.underline else False
                                code = font.name == "Consolas" if font and font.name else False
                                
                                val_str = wrap_text_style(val_str, bold=bold, italic=italic, strike=strike, underline=underline, code=code)
                            
                        if val_str and cell.hyperlink and cell.hyperlink.target:
                            val_str = f"[{val_str}]({cell.hyperlink.target})"
                            
                        val_str = val_str.replace("\n", " ").replace("|", "\\|")
                        row_cells.append(val_str)
                    grid.append(row_cells)
                    
                while grid and all(not c for c in grid[-1]):
                    grid.pop()
                    
                if not grid:
                    parts.append("*(Empty Table)*\n")
                    continue
                    
                header = "| " + " | ".join(grid[0]) + " |"
                sep = "| " + " | ".join("---" for _ in grid[0]) + " |"
                parts.append(header)
                parts.append(sep)
                for row in grid[1:]:
                    parts.append("| " + " | ".join(row) + " |")
                parts.append("")
                
            return "\n".join(parts)
        except Exception as e:
            import sys
            print(f"[DEBUG] Custom Excel parsing failed: {e}. Falling back to markitdown.", file=sys.stderr)
            try:
                from markitdown import MarkItDown
                md = MarkItDown()
                result = md.convert(file_path)
                if not result or not result.text_content:
                    return "*(Empty Excel Workbook)*"
                return result.text_content
            except Exception as inner_e:
                raise RuntimeError(f"Excel Ingestion Error: Failed to extract text layer from spreadsheet file. Detail: {str(inner_e)}")

    def save_from_markdown(self, markdown_content: str, out_path: str) -> str:
        """Converts Markdown line-by-line to a single Excel sheet, splitting tables to columns and keeping other text in Column A."""
        import re
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        # Set a clean default sheet title based on the output filename
        sheet_title = os.path.splitext(os.path.basename(out_path))[0][:31]
        sheet_title = re.sub(r'[\\/?*\[\]:]', "_", sheet_title)
        if not sheet_title:
            sheet_title = "Sheet1"
        ws.title = sheet_title

        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        body_font = Font(name="Arial", size=11)
        heading_font = Font(name="Arial", size=13, bold=True)
        thin = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def clean_md_inline(text: str) -> str:
            if not text:
                return ""
            # Clean HTML image tags
            text = re.sub(r'<p[^>]*?>\s*<img[^>]*?alt=["\']([^"\']*)["\'][^>]*?>\s*</p>', r'[\1]', text, flags=re.IGNORECASE)
            text = re.sub(r'<img[^>]*?alt=["\']([^"\']*)["\'][^>]*?>', r'[\1]', text, flags=re.IGNORECASE)
            text = re.sub(r'<img[^>]*?>', r'[Image]', text, flags=re.IGNORECASE)
            text = re.sub(r'!\[([^\]]*)\]\([^\)]+\)', r'[\1]', text)
            text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)
            text = re.sub(r'[*_~`]', '', text)
            if text.startswith(("- ", "* ")):
                text = "• " + text[2:]
            return text.strip()

        row_idx = 1
        in_table = False

        lines = markdown_content.splitlines()
        for line in lines:
            stripped = line.strip()
            if not stripped:
                in_table = False
                row_idx += 1
                continue

            # Skip table separator rows but ensure we are marked as in a table
            if "|" in stripped and re.match(r"^[\|\s\-:]+$", stripped):
                in_table = True
                continue

            # Check if it is a table row
            if "|" in stripped:
                inner_line = stripped
                if inner_line.startswith("|"):
                    inner_line = inner_line[1:]
                if inner_line.endswith("|"):
                    inner_line = inner_line[:-1]

                cells = [clean_md_inline(c) for c in inner_line.split("|")]
                is_header = not in_table
                in_table = True

                for col_idx, cell_text in enumerate(cells, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_text)
                    cell.border = thin_border
                    if is_header:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.font = body_font
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                row_idx += 1
            else:
                in_table = False
                cell = ws.cell(row=row_idx, column=1)

                match_heading = re.match(r"^(#{1,6})\s+(.*)", stripped)
                if match_heading:
                    heading_text = clean_md_inline(match_heading.group(2))
                    cell.value = heading_text
                    cell.font = heading_font
                else:
                    text_val = clean_md_inline(line)
                    cell.value = text_val
                    cell.font = body_font

                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                row_idx += 1

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            if max_len > 0:
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 5, 60)

        wb.save(out_path)
        return f"Exported 1 sheet successfully -> {os.path.basename(out_path)}"

ModuleRegistry.register(ExcelModule())
